import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.validator import EXPECTED_SCHEMA

def generate_template(file_path: str):
    """
    Generates a starter template with the expected structure and sample data.
    """
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Stylings
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Classic BPS blue
    centered_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # Mock data to help user understand format
    sample_data = {
        'data_petugas': [
            {
                'nama_bps': 'Ahmad', 'nik': '1234567890123456', 'nama_lengkap': 'Ahmad Fauzi', 
                'jabatan': 'PML', 'wilayah_tugas': 'Kec. Banjarmasin Tengah', 
                'custNoRef': 'REF001', 'no_rekening': '9876543210', 'is_bri': 'Ya', 'id_bank': '002', 
                'no_spk': '001/SPK/SE26/2026', 'no_bapp': '001/BAPP/SE26/2026', 'no_sppl': '001/SPPL/SE26/2026', 
                'bukti_bapp': 'Ada', 'tgl_penyelesaian_t1': '2026-08-25', 
                'min_jml_sls': '10', 'jml_sls': '15', 'listed_usaha_t1': '200', 'target_usaha_t1': '250'
            }
        ],
        'data_organik': [
            {
                'Nama': 'Budi Santoso', 'custNoRef': 'REF099', 'no_rekening': '1122334455', 
                'is_bri': 'Ya', 'id_bank': '002'
            }
        ],
        'merged': [
            {
                'nama_bps': 'Ahmad', 'custNoRef': 'REF001', 'no_rekening': '9876543210', 
                'is_bri': 'Ya', 'id_bank': '002', 'jabatan': 'PML'
            }
        ]
    }
    
    for sheet_name, cols in EXPECTED_SCHEMA.items():
        ws = wb.create_sheet(title=sheet_name)
        
        # Write headers
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_align
            cell.border = thin_border
            
        # Write sample row
        samples = sample_data.get(sheet_name, [])
        for row_idx, sample in enumerate(samples, 2):
            for col_idx, col_name in enumerate(cols, 1):
                val = sample.get(col_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = left_align
                cell.border = thin_border
                
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save the file
    wb.save(file_path)