"""Catalog-wide regression tests for Excel and generator routing."""
import unittest
import os
import tempfile
from pathlib import Path

import openpyxl
from docx import Document

from src import (
    bapp_pml,
    bapp_pml_t2,
    bapp_ppl,
    bapp_ppl_t2,
    bast,
    bukti_terima,
    lampiran_spk,
    spp,
    spp_t2,
)
from src.workflow import DOCUMENT_TYPES, get_document_by_id
from src.workflow_routing import (
    get_document_generator,
    get_input_validator,
    validate_document_input,
)
from src.document_generator import extend_input_template


ROOT = Path(__file__).resolve().parents[1]


class WorkflowRoutingTests(unittest.TestCase):
    EXPECTED_GENERATORS = {
        "lampiran_spk_ppl": lampiran_spk.iter_generate,
        "lampiran_spk_pml": lampiran_spk.iter_generate,
        "bapp_ppl_t1": bapp_ppl.iter_generate,
        "bapp_pml_t1": bapp_pml.iter_generate,
        "spp_ppl": spp.iter_generate,
        "spp_pml": spp.iter_generate,
        "bapp_ppl_t2": bapp_ppl_t2.iter_generate,
        "bapp_pml_t2": bapp_pml_t2.iter_generate,
        "spp_t2_ppl": spp_t2.iter_generate,
        "spp_t2_pml": spp_t2.iter_generate,
        "bast_ppl": bast.iter_generate,
        "bast_pml": bast.iter_generate,
        "bukti_terima": bukti_terima.iter_generate,
    }

    def test_every_catalog_document_has_validator_generator_and_assets(self):
        for document in DOCUMENT_TYPES:
            with self.subTest(document=document.id):
                self.assertIsNotNone(get_input_validator(document))
                self.assertIsNotNone(get_document_generator(document))
                self.assertIs(
                    get_document_generator(document),
                    self.EXPECTED_GENERATORS[document.id],
                )
                self.assertIsNotNone(document.input_template_path)
                self.assertTrue(Path(document.input_template_path).is_file())
                if not document.no_template:
                    self.assertIsNotNone(document.builtin_template_path)
                    self.assertTrue(Path(document.builtin_template_path).is_file())

    def test_every_bundled_input_validates_through_ui_routing(self):
        for document in DOCUMENT_TYPES:
            with self.subTest(document=document.id):
                template = None if document.no_template else document.builtin_template_path
                ok, errors, dfs = validate_document_input(
                    document,
                    document.input_template_path,
                    template_path=template,
                )
                self.assertTrue(ok, errors)
                self.assertTrue(dfs)

    def test_termin_1_workbook_reports_termin_2_specific_column(self):
        termin_1 = get_document_by_id("spp_ppl")
        termin_2 = get_document_by_id("spp_t2_ppl")

        ok, errors, _ = validate_document_input(
            termin_2,
            termin_1.input_template_path,
            template_path=termin_2.builtin_template_path,
        )

        self.assertFalse(ok)
        self.assertIn("no_urut_spp_t2", errors[0])

    def test_every_bapp_spp_and_bast_variant_requires_its_custom_column(self):
        target_ids = [
            document.id for document in DOCUMENT_TYPES
            if document.id.startswith(("bapp_", "spp_", "bast_"))
        ]
        self.assertEqual(len(target_ids), 10)

        for document_id in target_ids:
            document_type = get_document_by_id(document_id)
            with self.subTest(document=document_id), tempfile.TemporaryDirectory() as directory:
                template_path = os.path.join(directory, "custom.docx")
                template = Document(document_type.builtin_template_path)
                template.add_paragraph("{{catalog_custom}}")
                template.save(template_path)

                ok, errors, _ = validate_document_input(
                    document_type, document_type.input_template_path,
                    template_path=template_path,
                )
                self.assertFalse(ok)
                self.assertIn("catalog_custom", errors[0])

                input_path = os.path.join(directory, "custom.xlsx")
                sheet_name = (
                    "input" if document_id.startswith("bapp_") else "data_mitra"
                )
                custom_fields = ["catalog_custom", "catalog_custom_second"]
                source_workbook = openpyxl.load_workbook(
                    document_type.input_template_path, read_only=True,
                )
                source_headers = [
                    cell.value for cell in next(
                        source_workbook[sheet_name].iter_rows()
                    )
                ]
                source_workbook.close()
                last_builtin_column = max(
                    index for index, value in enumerate(source_headers, start=1)
                    if value is not None
                )

                extend_input_template(
                    document_type.input_template_path, input_path,
                    sheet_name, custom_fields,
                )
                generated_workbook = openpyxl.load_workbook(
                    input_path, read_only=True,
                )
                generated_sheet = generated_workbook[sheet_name]
                self.assertEqual(
                    generated_sheet.cell(1, last_builtin_column + 1).value,
                    "catalog_custom",
                )
                self.assertEqual(
                    generated_sheet.cell(1, last_builtin_column + 2).value,
                    "catalog_custom_second",
                )
                generated_workbook.close()
                ok, errors, _ = validate_document_input(
                    document_type, input_path, template_path=template_path,
                )
                self.assertTrue(ok, errors)

    def test_bundled_spp_assets_use_current_input_names_and_layout(self):
        termin_1_path = ROOT / "input" / "02_input_spp_t1.xlsx"
        workbook = openpyxl.load_workbook(termin_1_path, read_only=True)
        headers = [
            cell.value for cell in next(workbook["no_spk"].iter_rows())
        ]
        workbook.close()
        self.assertIn("no_urut_spp_t1", headers)
        self.assertNotIn("no_input_spp_t1", headers)

        termin_2_path = ROOT / "input" / "04_input_spp_t2.xlsx"
        workbook = openpyxl.load_workbook(termin_2_path, read_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["data_mitra", "no_spk", "alokasi_usaha"],
        )
        headers = [
            cell.value for cell in next(workbook["no_spk"].iter_rows())
        ]
        workbook.close()
        self.assertIn("no_urut_spp_t2", headers)
        self.assertNotIn("no_input_spp_t2", headers)
        self.assertNotIn("no_input_spp_t1", headers)

        for document_id in ("spp_t2_ppl", "spp_t2_pml"):
            fields = spp_t2._template_fields(
                get_document_by_id(document_id).builtin_template_path
            )
            self.assertIn("no_urut_spp_t2", fields)
            self.assertNotIn("no_input_spp_t2", fields)


if __name__ == "__main__":
    unittest.main()